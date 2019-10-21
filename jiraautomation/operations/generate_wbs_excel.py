from jiraautomation.operations.operation import basic_operation
from openpyxl import load_workbook

from .generate_wbs_json import generate_wbs_json
import pandas as pd
import numpy as np
import yaml


class generate_wbs_excel(basic_operation):

    @staticmethod
    def name():
        return "GenerateWBSExcel"

    @staticmethod
    def init_arguments(operation_group):
        generate_wbs_json.init_arguments(operation_group)
        operation_group.add_argument('-gwbss', '--generatewbs_sprints2fcs', required=False,
                                     help='Path to YAML file containing fcs and its corresponding sprints')

    @staticmethod
    def parse_arguments(args):
        generate_wbs_json.parse_arguments(args)
        pass

    def __init__(self, iLogger):
        super(generate_wbs_excel, self).__init__(iLogger)

    def execute(self, container, args):
        l = self.logger

        try:
            jira = container.getJIRA()

            try:
                op = generate_wbs_json(l)
                json_data = op.execute(container, args)
                df = pd.read_json(json_data)

                with open(args.generatewbs_sprints2fcs) as f:
                    fcs_sprints = yaml.load(f, Loader=yaml.Loader)

                expanded_df = calculated_cols(df)
                df_with_links = id_as_link(expanded_df)

                final_df = remove_pert_from_duplicates(df_with_links)

                fcs = sprints_mapping(final_df, fcs_sprints)

                save_data_to_excel(args.output, final_df, fcs)

            except Exception as e:
                l.error("Exception happened boards search " + str(e), e)

        except Exception as e:
            l.error("Exception happened during connection establishment " + str(e), e)


def calculated_cols(df):
    df['Original hours'] = df['Original est.'] / 3600
    df['Original days'] = df['Original hours'] / 8
    df['Spent hours'] = df['Spent time'] / 3600
    df['Spent days'] = df['Spent hours'] / 8
    df['Remaining hours'] = df['Remaining est.'] / 3600
    df['Remaining days'] = df['Remaining hours'] / 8

    return df


def sprints_mapping(df, FC_sprints):
    FC_reverse = dict((v, k) for k in FC_sprints for v in FC_sprints[k])

    df['Stages'] = df['Sprint'].map(FC_reverse)
    fcs = df['Stages'].dropna().unique()
    return fcs


def remove_pert_from_duplicates(df):
    df.loc[df.duplicated(subset=['ID']), ['PERT Opt', 'PERT Estimation (calculated)', "PERT Real",
                                          "PERT Pess", "PERT Estimation (calculated)"]] = np.nan
    return df


def id_as_link(df):
    df['ID'] = df['ID'].apply(lambda x: '=HYPERLINK("https://drivings.atlassian.net/browse/{}", "{}")'.format(x, x))
    return df

def save_data_to_excel(file, data, fcs):
    with pd.ExcelWriter(file, engine='openpyxl') as writer:
        writer.book = load_workbook(file)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        for fc in fcs:
            fc_to_pi = {'FC2': 'PI04', 'FC3': 'PI05', 'FC4': 'PI06'}
            if fc not in writer.book.sheetnames:
                sheet = fc_to_pi.get(fc)
            else:
                sheet = fc
            writer.sheet = writer.book.get_sheet_by_name(sheet)
            for row in writer.sheet:
                if row:
                    for cell in row:
                        cell.value = None

            fc_data = data[data['Stages'] == fc]
            fc_data = fc_data[fc_data.columns[:-1]]
            fc_data.to_excel(writer, sheet_name=sheet, index=False)
