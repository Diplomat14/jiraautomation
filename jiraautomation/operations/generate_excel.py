from jiraautomation.operations.operation import basic_operation
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
import pandas as pd


class generate_excel(basic_operation):

    @staticmethod
    def name():
        return "GenerateExcel"

    @staticmethod
    def init_arguments(operation_group):
        operation_group.add_argument('-genexc_data', '--generateexcel_Data', required=False,
                                     help='Input data to excel file')
        operation_group.add_argument('-genexc_sheet', '--generateexcel_SheetName', required=False,
                                     help='Name of excel sheet to fill out')
        operation_group.add_argument('-genexc_startcolumn', '--generateexcel_StartColumn', required=False,
                                     default=0,
                                     help='Initial column number', type=int)
        operation_group.add_argument('-genexc_startrow', '--generateexcel_StartRow', required=False, default=1,
                                     help='Initial row number', type=int)
        operation_group.add_argument('-genexc_index', '--generateexcel_Index', required=False, default=True,
                                     help='Flag if index should be added', type=bool)
        operation_group.add_argument('-genexc_columns', '--generateexcel_Columns', required=False,
                                     help='Сolumns numbers to convert to hyperlinks')
        operation_group.add_argument('-genexc_fontname', '--generateexcel_FontName', required=False,
                                     help='Font name for hyperlinks', default='Calibri')

    @staticmethod
    def parse_arguments(args):
        pass

    def __init__(self, iLogger):
        super(generate_excel, self).__init__(iLogger)

    def execute(self, container, args):
        l = self.logger

        try:

            try:
                df = pd.DataFrame(args.generateexcel_Data)
                if all(isinstance(clm, tuple) for clm in df.columns.values):
                    df.columns = pd.MultiIndex.from_tuples(df.columns)

                clms = []
                if args.generateexcel_Columns:
                    clms = [int(clm) for clm in args.generateexcel_Columns.split(',')]

                save_data_to_excel(args.output, df, args.generateexcel_SheetName, args.generateexcel_StartColumn,
                                   args.generateexcel_StartRow, args.generateexcel_Index,
                                   args.generateexcel_FontName, clms)

            except Exception as e:
                l.error("Exception happened boards search " + str(e), e)

        except Exception as e:
            l.error("Exception happened during connection establishment " + str(e), e)


def create_hyperlinks(writer, sheet_name, clms, font_name):
    ws = writer.book.get_sheet_by_name(sheet_name)
    for c in clms:
        for r in range(1, ws.max_row+1):
            if not ws.cell(row=r, column=c).value:
                continue

            cell_data = ws.cell(row=r, column=c).value.split(',')
            if len(cell_data) > 1:
                ws.cell(row=r, column=c).value = cell_data[1]
                ws.cell(row=r, column=c).hyperlink = cell_data[0]
                ws.cell(row=r, column=c).font = Font(u='single', color=colors.BLUE, size=10, name=font_name)


def save_data_to_excel(file, data, sheet_name, startcol, startrow, index, font_name, clms):
    with pd.ExcelWriter(file, engine='openpyxl') as writer:
        writer.book = load_workbook(file)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        data.to_excel(writer, sheet_name=sheet_name, startcol=startcol, startrow=startrow, index=index)
        if clms:
            create_hyperlinks(writer, sheet_name, clms, font_name)
